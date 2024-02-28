import ast
import os

import openpyxl
import pytest
from keys_all.web_keys.qiuyi_keys import *
from common.excel_handler import ExcelHandler




class Test003Qiuyi:
    root_path = os.path.dirname(__file__).split(r'\test_case')[0]
    test_path = root_path + r'/data/web_test_baidu.xlsx'

    @pytest.mark.demo
    @pytest.mark.parametrize('sheetnames', read_excel(test_path).sheetnames)
    def test_001_case(self,sheetnames):
        '''
        file为excel对象
        :param sheetnames: 这里传递进来得时str，应该是list才行

        :return:
        '''
        excel = read_excel(self.test_path)
        try:
            for name in sheetnames:
                sheet = excel[name]  # 通过sheet的name名字遍历sheet
                print('开始测试{}文件中的用例'.format(name).center(40, '>'))
                for values in sheet.values:  # 获取sheet中所有值
                    sheet.cell(row=123, column=43).coordinate
                    case_para = del_dict_None(values)
                    if type(case_para[0]) is int:  # 排除掉第一行
                        # 实例化对象
                        # self.log.info("用例：{},开始执行".format(values[3]))
                        keys = init_driver(case_para)
                        exec_case(case_para, keys, sheet, name, excel, self.test_path)
        except Exception as e:
            # self.log.Exception('运行异常：{}'.format(e))
            print(e)
        finally:
            excel.close()

if __name__ == '__main__':
    print(Test003Qiuyi().test_path)