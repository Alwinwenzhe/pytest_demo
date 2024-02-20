import ast, xlrd
from xlutils.copy import copy
from openpyxl.styles import PatternFill, Font
from data import Config
from common import operate_json

class ExcelHandler(object):
    '''
    关于Excel表的操作
    '''
    oper_j = operate_json.OperateJson()
    con = Config.Config()

    def read_excel(self,sheet_path):
        '''
        读取固定excel,返回指定sheet内容
        :return:
        '''
        book = xlrd.open_workbook(sheet_path)  # 注意：这里已经写死了是接口测试对应的文件路径，并没有参数化
        table = book.sheet_by_index(0)        # 通过索引顺序获取sheet
        return table

    def get_excel_data(self,sheet_path=Config.API_CASE_PATH):
        '''
        可以通过参数：file_name文件名来区分不同数据，对应不同的函数入口
         数据文件过滤excel中不必要的数据
        :param file_name: 通过他来区分是接口还是web 数据文件
        :param case_desc: 通过excel中的case_description来过滤用例
        :return:
        '''
        sheet = self.read_excel(sheet_path)
        rows, cols = sheet.nrows, sheet.ncols
        l = []
        title = sheet.row_values(0)
        # 获取其他行
        for i in range(1, rows):
            #print(sheet.row_values(i))
            # if case_desc in sheet.row_values(i):
            l.append(dict(zip(title, sheet.row_values(i))))  # 先返回一个zip对象，按最短得title或row_values拼接；然后通过dict格式化为dict，最后增加到list中
        return l

    def write_excel_by_xlutils(self,row,value):
        '''
        对第4列的值进行修改
        :param row: 从1开始
        :param value:
        :return:
        '''
        workbook = xlrd.open_workbook(Config.API_CASE_PATH)
        workbook_new= copy(workbook)
        sheet = workbook_new.get_sheet(0)

        print(sheet)
        # sheet.write(row,3,value)
        # workbook_new.save(Config.API_CASE_PATH)

    def write_excel(self,keys,sheet,values,name):
        '''
        写入数据到excel
        :param sheet: 当前sheet
        :param values: 读取得数据内容
        :param name: sheet得name
        :return:
        '''
        para = self.str_to_dict(values[2])
        result = getattr(keys, values[1])(**para)
        if result:
            print('Case：{}，执行成功'.format(name))
            # sheet.cell(values[0] + 1, column=5).value('Pass')
            self.write_result(sheet.cell,values[0]+1,5,'Pass')
        else:
            print('Case：{}，执行失败'.format(name))
            # sheet.cell(values[0] + 1, column=5).value('False')
            self.write_result(sheet.cell,values[0]+1,5,'Fail')

    def str_to_dict(self,str_data):
        '''
        将字符串格式得数据转化为python可识别得dict格式;比使用json.loads更好，因为str如果内容是单引号会报错
        :param str_data:
        :return:
        '''
        return ast.literal_eval(str_data)

    def write_result(self,cell,row,column,result):
        '''设置Pass样式'''
        if result == 'Pass':
            color = 'AACF91'        # 绿色
            # 写入内容
            cell(row,column).value = 'Pass'
        elif result == "Fail":
            color = 'FF0000'  # 红色
            # 写入内容
            cell(row, column).value = 'Fail'
        # 单元格样式定义：绿色+加粗
        cell(row, column).fill = PatternFill('solid', color)  # 设置颜色
        cell(row, column).font = Font(bold=True)  # 加粗

    def update_column_excel(self,key,add_cont=None,sheet_index=0,file_path=Config.API_CASE_PATH):
        '''
        按行批量修改特定某列数据,数据写入excel未完成
        PASS
        :param key:
        :param add_cont:     添加部分
        :param sheet_index:
        :return:
        '''
        row = 1
        workbook = xlrd.open_workbook(file_path)
        workbook_new = copy(workbook)
        sheet = workbook_new.get_sheet(sheet_index)
        excel_list = self.get_excel_data()
        for i in excel_list:
            if add_cont != None:
                i[key] = add_cont + i[key]
            sheet.write(row, 3, i[key])
            row += 1
        workbook_new.save(file_path)

    def update_id_card(self,key,sheet_index=0,file_path=Config.API_CASE_PATH):
        '''
        按行批量写入身份证号码
        :param key:
        :param add_cont:     添加部分
        :param sheet_index:
        :return:
        '''
        id_card = 513902198604102110
        row = 1
        workbook = xlrd.open_workbook(file_path)
        workbook_new = copy(workbook)
        sheet = workbook_new.get_sheet(sheet_index)
        excel_list = self.get_excel_data()
        for i in excel_list:

            i[key] = str(id_card + 1)
            sheet.write(row, 3, i[key])
            id_card +=1
            row += 1
        workbook_new.save(file_path)

    def replace_column_excel(self,key,replace_old=None,replace_new=None,replace_all=None,sheet_index=0,file_path=Config.API_CASE_PATH):
        '''
        按行批量修改特定某列数据,数据写入excel未完成
        PASS
        :param key:
        :param replace_old:    替换部分目标内容
        :param replace_new:    替换源
        :param replace_all:     替换所有
        :param sheet_index:
        :return:
        '''
        row = 1
        workbook = xlrd.open_workbook(file_path)
        workbook_new = copy(workbook)
        sheet = workbook_new.get_sheet(sheet_index)
        excel_list = self.get_excel_data()
        for i in excel_list:
            if replace_all !=None:
                i[key] = replace_all
            if replace_old != None:
                i[key] = i[key].replace(replace_old,replace_new)
            sheet.write(row, 3, i[key])
            row += 1
        workbook_new.save(file_path)

if __name__ == '__main__':
    eh = ExcelHandler()
    # eh.update_column_excel("name",add_cont="双语国际_",file_path=r'J:\now_job\西交智汇\data\学校\AT_高新-3\成华_双语国际-非\学生_500.xlsx')
    eh.update_id_card('id',file_path=r'J:\now_job\西交智汇\data\学校\AT_高新-3\成华_双语国际-非\学生_500.xlsx')